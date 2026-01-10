# --- Static Template Preview Views ---
from django.views.generic import TemplateView





from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth.hashers import make_password
import openpyxl
from openpyxl.styles import Font
from django.urls import reverse
from .utils import create_notification  

from .models import CustomUser
from employee_app.permission_defaults import ROLE_PERMISSIONS
from employee_app.utils import create_notification, has_permission

def login_view(request):
    if request.user.is_authenticated:
        if request.user.role == 'employee':
            return redirect('employee_app:employee_dashboard')
        return redirect('employee_app:users_manager')

    if request.method == 'POST':
        email = request.POST.get('email').lower().strip()
        password = request.POST.get('password')

        user = authenticate(request, email=email, password=password)

        if user is not None:
            if user.status == 'active':
                login(request, user)
                return redirect('employee_app:dashboard')
            else:
                messages.error(request, 'Account is inactive.')
        else:
            messages.error(request, 'Invalid email or password.')

    return render(request, 'employee_app/login.html')


def logout_view(request):
    logout(request)
    messages.success(request, 'Logged out successfully.')
    return redirect('employee_app:login')


@login_required
def users_manager(request):
    if not has_permission(request.user, 'users', 'view'):
        return HttpResponse('Access Denied', status=403)

    users = CustomUser.objects.all().order_by('-date_joined')
    return render(request, 'employee_app/users_manager.html', {'users': users})





@login_required
def create_user(request):
    if not has_permission(request.user, 'users', 'create'):
        return HttpResponse('Access Denied', status=403)

    if request.method == 'POST':
        email = request.POST.get('email').lower().strip()

        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists.')
            return redirect('employee_app:users_manager')

        role = request.POST.get('role')
        permissions = ROLE_PERMISSIONS.get(role, {})

        new_user = CustomUser.objects.create(
            email=email,
            full_name=request.POST.get('full_name'),
            phone=request.POST.get('phone'),
            department=request.POST.get('department') or '',
            role=role,
            status=request.POST.get('status'),
            permissions=permissions,
            password=make_password(request.POST.get('password')),
            is_active=(request.POST.get('status') == 'active')
        )

        # If employee, optional biodata link
        if role == 'employee':
            bio_id = request.POST.get('bio_data')
            if bio_id:
                try:
                    bio = BioDataRequest.objects.get(id=bio_id, status='approved')
                    if not bio.user:
                        bio.user = new_user
                        bio.save()
                        messages.info(request, "Linked to selected biodata.")
                    else:
                        messages.warning(request, "Biodata already linked to another user.")
                except BioDataRequest.DoesNotExist:
                    messages.warning(request, "Selected biodata not found or not approved.")

        # Notification
        for notifier in CustomUser.objects.filter(role__in=['admin', 'super_admin']):
            create_notification(
                notifier,
                'user_created',
                'New User Created',
                f"{new_user.email} ({new_user.get_role_display()}) created by {request.user.email}.",
            )

        messages.success(request, f'User {email} created successfully.')
        return redirect('employee_app:users_manager')

    # For GET: pass approved biodata for dropdown
    available_biodata = BioDataRequest.objects.filter(status='approved', user__isnull=True)
    return render(request, 'employee_app/create_user.html', {'available_biodata': available_biodata})



from .forms import UserEditForm
@login_required
def edit_user(request, user_id):
    if not has_permission(request.user, 'users', 'edit'):
        return HttpResponse('Access Denied', status=403)

    user = get_object_or_404(CustomUser, id=user_id)

    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()

            # Reset permissions if role changed
            user.permissions = ROLE_PERMISSIONS.get(user.role, {})
            user.save()

            # Notification
            create_notification(
                recipient=user,
                ntype='user_updated',
                title='Your Profile Was Updated',
                message=f"Your account details were updated by {request.user.get_full_name() or request.user.email}.",
                link=request.build_absolute_uri(reverse('employee_app:profile'))
            )

            messages.success(request, 'User updated successfully.')
            return redirect('employee_app:users_manager')
    else:
        form = UserEditForm(instance=user)

    return render(request, 'employee_app/edit_user.html', {'form': form, 'user': user})


@login_required
def delete_user(request, user_id):
    if not has_permission(request.user, 'users', 'delete'):
        return HttpResponse('Access Denied', status=403)

    if request.method == 'POST':
        user = get_object_or_404(CustomUser, id=user_id)

        if user == request.user:
            messages.error(request, "You cannot delete your own account.")
        else:
            deleted_email = user.email
            deleted_role = user.get_role_display()
            user.delete()

            # === NOTIFICATION: User deleted → notify admin & super_admin ===
            for notifier in CustomUser.objects.filter(role__in=['admin', 'super_admin'], is_active=True):
                create_notification(
                    recipient=notifier,
                    ntype='user_deleted',
                    title='User Deleted',
                    message=f"User {deleted_email} ({deleted_role}) was deleted by {request.user.get_full_name() or request.user.email}.",
                )

            messages.success(request, "User deleted successfully.")

    return redirect('employee_app:users_manager')

@login_required
def export_users_excel(request):
    if not has_permission(request.user, 'users', 'export'):
        return HttpResponse('Access Denied', status=403)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="users.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    columns = ['Full Name', 'Email', 'Phone', 'Department', 'Role', 'Status', 'Date Joined']
    ws.append(columns)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for user in CustomUser.objects.all():
        ws.append([
            user.full_name or '-',
            user.email,
            user.phone or '-',
            user.get_department_display(),
            user.get_role_display(),
            user.get_status_display(),
            user.date_joined.strftime('%Y-%m-%d')
        ])

    wb.save(response)
    return response

@login_required
def dashboard(request):
    if not has_permission(request.user, 'dashboard', 'view'):
        return HttpResponse('Access Denied', status=403)
    return render(request, 'employee_app/dashboard.html')






# @login_required
# def training(request):
#     if not has_permission(request.user, 'training', 'view'):
#         return HttpResponse('Access Denied', status=403)
#     return render(request, 'employee_app/training.html')

# @login_required
# def add_trainee(request):
#     if not has_permission(request.user, 'training', 'create'):
#         return HttpResponse('Access Denied', status=403)
#     return render(request, 'employee_app/add_trainee.html')



@login_required
def projects(request):
    if not has_permission(request.user, 'projects', 'view'):
        return HttpResponse('Access Denied', status=403)
    return render(request, 'employee_app/projects.html')



@login_required
def profile(request):
    return render(request, 'employee_app/profile.html')

# def public_biodata_form(request):
#     return render(request, 'employee_app/public_biodata_form.html') 

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.sessions.models import Session
from django.shortcuts import render, redirect
from django.utils import timezone
from django.views.decorators.http import require_POST


@login_required
def app_settings(request):
    sessions = Session.objects.filter(expire_date__gte=timezone.now())
    user_sessions = []

    for session in sessions:
        data = session.get_decoded()
        if data.get('_auth_user_id') == str(request.user.id):
            user_sessions.append(session)

    return render(request, 'employee_app/settings.html', {
        'active_sessions': user_sessions
    })


@login_required
@require_POST
def change_password(request):
    user = request.user

    current_password = request.POST.get('current_password')
    new_password = request.POST.get('new_password')
    confirm_password = request.POST.get('confirm_password')

    if not user.check_password(current_password):
        messages.error(request, "Current password is incorrect")
        return redirect('employee_app:settings')

    if new_password != confirm_password:
        messages.error(request, "Passwords do not match")
        return redirect('employee_app:settings')

    if len(new_password) < 8:
        messages.error(request, "Password must be at least 8 characters")
        return redirect('employee_app:settings')

    user.set_password(new_password)
    user.save()
    update_session_auth_hash(request, user)

    messages.success(request, "Password changed successfully")
    return redirect('employee_app:settings')


# @login_required
# @require_POST
# def update_preferences(request):
#     user = request.user
#     user.email_notifications = request.POST.get('email_notifications')
#     user.timezone = request.POST.get('timezone')
#     user.save()

#     messages.success(request, "Preferences updated successfully")
#     return redirect('employee_app:settings')


@login_required
@require_POST
def sign_out_all_devices(request):
    sessions = Session.objects.filter(expire_date__gte=timezone.now())

    for session in sessions:
        data = session.get_decoded()
        if data.get('_auth_user_id') == str(request.user.id):
            session.delete()

    messages.success(request, "Signed out from all devices")
    return redirect('employee_app:settings')



# Public Form
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth.hashers import make_password
import openpyxl
from openpyxl.styles import Font

from .models import CustomUser, BioDataRequest
from .forms import BioDataForm, ReviewForm
from employee_app.permission_defaults import ROLE_PERMISSIONS
from employee_app.utils import has_permission
from django.core.mail import send_mail
from django.conf import settings

def public_biodata_form(request):
    if request.method == 'POST':
        form = BioDataForm(request.POST, request.FILES)
        if form.is_valid():
            work_exp = []
            employers = request.POST.getlist('prev_employer[]')
            designations = request.POST.getlist('prev_designation[]')
            durations = request.POST.getlist('prev_duration[]')
            emails = request.POST.getlist('prev_email[]')
            cert_files = request.FILES.getlist('work_experience_cert[]')

            for i in range(len(employers)):
                if employers[i].strip():
                    exp = {
                        'employer': employers[i],
                        'designation': designations[i],
                        'duration': durations[i],
                        'email': emails[i],
                    }

                    if i < len(cert_files) and cert_files[i]:
                        cert_file = cert_files[i]
                        from django.core.files.storage import default_storage
                        path = default_storage.save(
                            f'biodata/certs/{cert_file.name}',
                            cert_file
                        )

                        # ✅ ONLY FIX IS HERE
                        exp['certificate_path'] = default_storage.url(path)

                    work_exp.append(exp)

            bio = form.save(commit=False)
            bio.work_experience = work_exp
            bio.save()

            # Notify admin and super_admin about new submission
            for user in CustomUser.objects.filter(role__in=['admin', 'super_admin'], is_active=True):
                link = reverse('employee_app:review_biodata_detail', args=[bio.pk])
                create_notification(
                    recipient=user,
                    ntype='biodata_new',
                    title='New BioData Submission',
                    message=f"{bio.first_name} {bio.last_name} submitted a new bio data request.",
                    link=request.build_absolute_uri(link)
                )

            messages.success(
                request,
                'Bio data submitted successfully! Awaiting HR review.'
            )
            return redirect('employee_app:public_biodata_form')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = BioDataForm()

    return render(
        request,
        'employee_app/public_biodata_form.html',
        {'form': form}
    )


@login_required
def pending_requests(request):
    requests = BioDataRequest.objects.all().order_by('-created_at')
    return render(request, 'employee_app/pending_requests.html', {'requests': requests})


import secrets
import string
@login_required
def review_biodata_detail(request, pk):
    bio = get_object_or_404(BioDataRequest, pk=pk)

    if request.method == 'POST':
        form = ReviewForm(request.POST, instance=bio)
        if form.is_valid():
            action = request.POST.get('action')
            create_account = form.cleaned_data.get('create_account', True)

            bio = form.save(commit=False)

            if action == 'approve':
                bio.status = 'approved'
                bio.approved_by = request.user

                if create_account and not bio.user:
                    email = bio.official_email or bio.personal_email
                    if CustomUser.objects.filter(email=email).exists():
                        messages.error(request, f"Email {email} already used.")
                    else:
                        temp_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(10))
                        new_user = CustomUser.objects.create(
                            email=email,
                            full_name=f"{bio.first_name} {bio.middle_name or ''} {bio.last_name}".strip(),
                            phone=bio.contact_number,  # COPY PHONE
                            department=bio.department,  # COPY DEPARTMENT
                            role='employee',
                            status='active',
                            password=make_password(temp_password),
                            is_active=True
                        )
                        bio.user = new_user
                        bio.save()

                       
                        # Send welcome email
                        subject = 'Welcome to STACKLY - Your Account Details'
                        message = f"""
                        Dear {bio.first_name} {bio.last_name},

                        Your bio data has been approved!

                        Login Details:
                        - Email: {email}
                        - Temporary Password: {temp_password}

                        Please login at: {request.build_absolute_uri(reverse('employee_app:login'))}
                        Change your password immediately after login.

                        Employee Details:
                        - Employee ID: {bio.employee_id}
                        - Official Email: {bio.official_email}
                        - Designation: {bio.designation}
                        - Department: {bio.department}
                        - Date of Joining: {bio.doj}

                        Best regards,
                        HR Team - STACKLY
                        """
                        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [bio.personal_email])

                        # Notify admins
                        for notifier in CustomUser.objects.filter(role__in=['admin', 'super_admin']):
                            create_notification(
                                notifier,
                                'employee_account_created',
                                'Employee Account Created',
                                f"Account created for {bio.first_name} {bio.last_name} ({email}) on approval.",
                                reverse('employee_app:edit_user', args=[new_user.id])
                            )

                messages.success(request, 'Employee approved successfully!' + (' Account created.' if create_account else ''))
            elif action == 'reject':
                bio.status = 'rejected'
                messages.success(request, 'Application rejected successfully.')

            bio.save()
            return redirect('employee_app:pending_requests')

    else:
        form = ReviewForm(instance=bio)

    return render(request, 'employee_app/review_biodata_detail.html', {'bio': bio, 'form': form})

from django.db.models import Q

@login_required
def biodata_list(request):
    employees = BioDataRequest.objects.filter(status='approved').order_by('-doj')

    # Get filter parameters from URL
    search = request.GET.get('search', '').strip()
    department = request.GET.get('department', '').strip()

    # Apply search filter (name, email, employee_id)
    if search:
        employees = employees.filter(
            Q(first_name__icontains=search) |
            Q(middle_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(employee_id__icontains=search) |
            Q(official_email__icontains=search) |
            Q(personal_email__icontains=search)
        )

    # Apply department filter
    if department:
        employees = employees.filter(department=department)

    context = {
        'employees': employees,
        'current_search': search,           # to keep input filled
        'current_department': department,   # to keep dropdown selected
    }
    return render(request, 'employee_app/biodata.html', context)

@login_required
def delete_pending_request(request, pk):
    if not has_permission(request.user, 'biodata', 'delete'):
        return HttpResponse('Access Denied', status=403)

    request_obj = get_object_or_404(BioDataRequest, pk=pk, status='pending')

    if request.method == 'POST':
        name = f"{request_obj.first_name} {request_obj.last_name}"
        request_obj.delete()
        messages.success(request, f"Pending request for {name} deleted.")
        return redirect('employee_app:pending_requests')

    return render(request, 'employee_app/delete_confirm.html', {
        'object': request_obj,
        'title': 'Delete Pending Request',
        'back_url': reverse('employee_app:pending_requests')
    })


@login_required
def delete_approved_employee(request, pk):
    if not has_permission(request.user, 'biodata', 'delete'):
        return HttpResponse('Access Denied', status=403)

    employee = get_object_or_404(BioDataRequest, pk=pk, status='approved')

    if request.method == 'POST':
        name = f"{employee.first_name} {employee.last_name}"
        employee.delete()
        messages.success(request, f"Approved employee {name} deleted.")
        return redirect('employee_app:biodata_list')

    return render(request, 'employee_app/delete_confirm.html', {
        'object': employee,
        'title': 'Delete Approved Employee',
        'back_url': reverse('employee_app:biodata_list')
    })

@login_required
def view_biodata(request, pk):
    bio = get_object_or_404(BioDataRequest, pk=pk, status='approved')
    return render(request, 'employee_app/view_biodata.html', {'bio': bio})

from .forms import BioDataEditForm
@login_required
def edit_biodata(request, pk):
    bio = get_object_or_404(BioDataRequest, pk=pk, status='approved')

    if request.method == 'POST':
        form = BioDataEditForm(request.POST, request.FILES, instance=bio)
        if form.is_valid():
            form.save()

            # Notification
            for notifier in CustomUser.objects.filter(role__in=['admin', 'super_admin'], is_active=True):
                link = reverse('employee_app:view_biodata', args=[bio.pk])
                create_notification(
                    recipient=notifier,
                    ntype='biodata_updated',
                    title='Employee BioData Updated',
                    message=f"BioData of {bio.first_name} {bio.last_name} (ID: {bio.employee_id or 'N/A'}) was updated by {request.user.get_full_name() or request.user.email}.",
                    link=request.build_absolute_uri(link)
                )

            messages.success(request, 'Employee bio data updated successfully!')
            return redirect('employee_app:view_biodata', pk=bio.pk)
    else:
        form = BioDataEditForm(instance=bio)

    return render(request, 'employee_app/edit_biodata.html', {'form': form, 'bio': bio})


@login_required
def delete_biodata(request, pk):
    if not has_permission(request.user, 'biodata', 'delete'):
        return HttpResponse('Access Denied', status=403)

    bio = get_object_or_404(BioDataRequest, pk=pk, status='approved')

    if request.method == 'POST':
        deleted_name = f"{bio.first_name} {bio.last_name}"
        deleted_email = bio.personal_email
        bio.delete()

        # === NOTIFICATION: BioData deleted → notify admin & super_admin ===
        for notifier in CustomUser.objects.filter(role__in=['admin', 'super_admin'], is_active=True):
            create_notification(
                recipient=notifier,
                ntype='biodata_deleted',
                title='Employee BioData Deleted',
                message=f"BioData of {deleted_name} ({deleted_email}) was deleted by {request.user.get_full_name() or request.user.email}.",
            )

        messages.success(request, f"BioData of {deleted_name} deleted successfully.")
        return redirect('employee_app:biodata_list')

    # If GET request → show confirmation page (optional but recommended)
    return render(request, 'employee_app/delete_biodata_confirm.html', {'bio': bio})


from datetime import datetime
from openpyxl.styles import Font, Alignment
@login_required
def export_biodata_excel(request):
    if not has_permission(request.user, 'biodata', 'export'):
        return HttpResponse('Access Denied', status=403)

    employees = BioDataRequest.objects.filter(status='approved').order_by('-doj')

    # Apply the same filters as the list view
    search = request.GET.get('search', '').strip()
    department = request.GET.get('department', '').strip()

    if search:
        employees = employees.filter(
            Q(first_name__icontains=search) |
            Q(middle_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(employee_id__icontains=search) |
            Q(official_email__icontains=search) |
            Q(personal_email__icontains=search)
        )

    if department:
        employees = employees.filter(department=department)

    # Now build Excel with filtered employees
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approved Employees"

    columns = [
        'ID', 'First Name', 'Middle Name', 'Last Name', 'Personal Email', 'Contact Number',
        'Employee ID', 'Official Email', 'Designation', 'Department', 'DOJ', 'Work Mode',
        'Experience Type', 'Post Applied For', 'Blood Group',
        'Address', 'Aadhar No', 'PAN No',
        'Bank Name', 'Branch', 'Account No', 'Account Name', 'IFSC',
        'Technical Skills', 'Created At'
    ]
    ws.append(columns)

    for col_num, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for emp in employees:
        address = f"{emp.address_line1 or ''} {emp.address_line2 or ''}, {emp.city or ''}, {emp.state or ''} {emp.postal_code or ''}, {emp.get_country_display() or ''}".strip()
        ws.append([
            emp.id,
            emp.first_name,
            emp.middle_name or '-',
            emp.last_name,
            emp.personal_email,
            emp.contact_number,
            emp.employee_id or '-',
            emp.official_email or '-',
            emp.designation or '-',
            emp.department or '-',
            emp.doj.strftime('%Y-%m-%d') if emp.doj else '-',
            emp.work_mode or '-',
            emp.get_experience_type_display(),
            emp.get_post_applied_for_display() or '-',
            emp.blood_group or '-',
            address or '-',
            emp.aadhar_no or '-',
            emp.pan_no or '-',
            emp.bank_name or '-',
            emp.bank_branch or '-',
            emp.account_number or '-',
            emp.account_name or '-',
            emp.ifsc_code or '-',
            emp.technical_skills,
            emp.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    # Auto width + freeze header (same as before)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
    ws.freeze_panes = 'A2'

    today = datetime.now().strftime('%Y-%m-%d')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="approved_employees_{today}.xlsx"'
    wb.save(response)
    return response



from django.http import JsonResponse

@login_required
@require_POST
def mark_all_read(request):
    request.user.notifications.update(is_read=True)
    return JsonResponse({'status': 'ok'})

@login_required
def notifications_all(request):
    notifs = request.user.notifications.all()
    return render(request, 'employee_app/notifications_all.html', {'notifications': notifs})

from .models import CustomUser, BioDataRequest, Notification
from .forms import BioDataForm, ReviewForm,  EmployeeProfileForm
from .utils import create_notification, has_permission

@login_required
def employee_dashboard(request):
    if request.user.role != 'employee':
        return redirect('employee_app:dashboard')  # redirect admins to main dashboard

    bio = request.user.bio_data_request if hasattr(request.user, 'bio_data_request') else None
    return render(request, 'employee_app/employee_dashboard.html', {'bio': bio})


# NEW: My Profile (view)
@login_required
def my_profile(request):
    if request.user.role != 'employee':
        return redirect('employee_app:dashboard')

    bio = request.user.bio_data_request
    if not bio:
        messages.warning(request, "No biodata linked to your account yet.")
        return redirect('employee_app:employee_dashboard')

    return render(request, 'employee_app/my_profile.html', {'bio': bio})


# NEW: Edit My Profile (limited fields)
@login_required
def edit_my_profile(request):
    if request.user.role != 'employee':
        return redirect('employee_app:dashboard')

    bio = request.user.bio_data_request
    if not bio:
        messages.error(request, "No biodata found.")
        return redirect('employee_app:my_profile')

    if request.method == 'POST':
        form = EmployeeProfileForm(request.POST, instance=bio)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your profile updated successfully!')
            return redirect('employee_app:my_profile')
    else:
        form = EmployeeProfileForm(instance=bio)

    return render(request, 'employee_app/edit_my_profile.html', {'form': form, 'bio': bio})








from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from django.utils import timezone
from django.http import HttpResponseForbidden
from .models import Batch, Session, Assignment, Submission
from .forms import (
  
    SessionForm, SessionEditForm,
    AssignmentForm, AssignmentEditForm,
    SubmissionForm
)

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q

@login_required
def training_dashboard(request):
    if not has_permission(request.user, 'training', 'view') or not has_permission(request.user, 'training', 'manage'):
        return HttpResponseForbidden("You don't have permission to access the training management dashboard.")

    # Search query from GET param (?q=...)
    search_query = request.GET.get('q', '').strip()

    # ───── Batches ─────
    batches_qs = Batch.objects.all().order_by('-start_date')
    if search_query:
        batches_qs = batches_qs.filter(
            Q(name__icontains=search_query) |
            Q(batch_code__icontains=search_query)
        )
    batches_paginator = Paginator(batches_qs, 10)
    batches_page = request.GET.get('batches_page', 1)
    try:
        batches = batches_paginator.page(batches_page)
    except PageNotAnInteger:
        batches = batches_paginator.page(1)
    except EmptyPage:
        batches = batches_paginator.page(batches_paginator.num_pages)

    # ───── Sessions ───── (newest first)
    sessions_qs = Session.objects.all().order_by('-date_time')
    if search_query:
        sessions_qs = sessions_qs.filter(
            Q(title__icontains=search_query) |
            Q(batch__name__icontains=search_query)
        )
    sessions_paginator = Paginator(sessions_qs, 10)
    sessions_page = request.GET.get('sessions_page', 1)
    try:
        sessions = sessions_paginator.page(sessions_page)
    except PageNotAnInteger:
        sessions = sessions_paginator.page(1)
    except EmptyPage:
        sessions = sessions_paginator.page(sessions_paginator.num_pages)

    # ───── Assignments ───── (newest due date first)
    assignments_qs = Assignment.objects.all().order_by('-due_date')
    if search_query:
        assignments_qs = assignments_qs.filter(
            Q(title__icontains=search_query) |
            Q(batch__name__icontains=search_query)
        )
    assignments_paginator = Paginator(assignments_qs, 10)
    assignments_page = request.GET.get('assignments_page', 1)
    try:
        assignments = assignments_paginator.page(assignments_page)
    except PageNotAnInteger:
        assignments = assignments_paginator.page(1)
    except EmptyPage:
        assignments = assignments_paginator.page(assignments_paginator.num_pages)

    context = {
        'batches': batches,
        'sessions': sessions,
        'assignments': assignments,
        'is_management_view': True,
        'search_query': search_query,  # to show in search box
        # Quick stats (unchanged)
        'active_batches_count': Batch.objects.filter(status='ongoing').count(),
        'total_sessions': Session.objects.count(),
        'pending_assignments': Assignment.objects.filter(status='pending').count(),
        'enrolled_employees': CustomUser.objects.filter(role='employee', enrolled_batches__isnull=False).distinct().count(),
    }

    return render(request, 'employee_app/training.html', context)

@login_required
def my_training(request):
    """
    Personal training dashboard for employees.
    Shows only their enrolled batches, upcoming sessions, pending assignments.
    """
    if request.user.role != 'employee':
        # Redirect admins/trainers to management dashboard
        return redirect('employee_app:training_admin_dashboard')

    my_batches = Batch.objects.filter(employees=request.user).order_by('-start_date')
    upcoming_sessions = Session.objects.filter(
        batch__employees=request.user,
        date_time__gte=timezone.now()
    ).order_by('date_time')[:5]

    pending_assignments = Assignment.objects.filter(
        batch__employees=request.user,
        status='pending'
    ).order_by('due_date')

    context = {
        'my_batches': my_batches,
        'upcoming_sessions': upcoming_sessions,
        'pending_assignments': pending_assignments,
        'is_employee_view': True,
    }

    return render(request, 'employee_app/my_training.html', context)


# ───────────────────────────────────────────────
# Batch Views
# ───────────────────────────────────────────────
# views.py
from .forms import BatchCreateForm , BatchUpdateForm
@login_required
def create_batch(request):
    if not has_permission(request.user, 'training', 'create'):
        return HttpResponseForbidden()

    if request.method == 'POST':
        form = BatchCreateForm(request.POST)
        if form.is_valid():
            batch = form.save(commit=False)
            batch.created_by = request.user
            batch.save()
            form.save_m2m()
            messages.success(request, f'Batch "{batch.name}" created successfully!')
            return redirect('employee_app:training_dashboard')
    else:
        form = BatchCreateForm()

    return render(request, 'employee_app/add_batch.html', {'form': form})


@login_required
def edit_batch(request, pk):
    if not has_permission(request.user, 'training', 'edit'):
        return HttpResponseForbidden()

    batch = get_object_or_404(Batch, pk=pk)

    if request.method == 'POST':
        form = BatchUpdateForm(request.POST, instance=batch)
        if form.is_valid():
            form.save()   # ← saves both fields + m2m relations
            messages.success(request, f'Batch "{batch.name}" updated successfully!')
            return redirect('employee_app:batch_detail', pk=batch.pk)
    else:
        form = BatchUpdateForm(instance=batch)

    return render(request, 'employee_app/edit_batch.html', {
        'form': form,
        'batch': batch  # ← useful for showing current count, etc.
    })


@login_required
def batch_detail(request, pk):
    batch = get_object_or_404(Batch, pk=pk)

    # Calculate duration in days
    if batch.start_date and batch.end_date:
        duration_days = (batch.end_date - batch.start_date).days
    else:
        duration_days = 0

    context = {
        'batch': batch,
        'sessions': batch.sessions.all().order_by('date_time'),
        'assignments': batch.assignments.all().order_by('due_date'),
        'employee_count': batch.employees.count(),
        'trainer_count': batch.trainers.count(),
        'duration_days': duration_days,  # ← New context variable
    }

    return render(request, 'employee_app/batch_detail.html', context)


# ───────────────────────────────────────────────
# Session Views
# ───────────────────────────────────────────────

@login_required
def create_session(request, batch_id=None):  # Add optional batch_id parameter
    if not has_permission(request.user, 'training', 'create'):
        return HttpResponseForbidden("You don't have permission to create sessions.")

    initial = {}
    if batch_id:
        initial['batch'] = batch_id

    if request.method == 'POST':
        form = SessionForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.created_by = request.user
            session.save()
            messages.success(request, f'Session "{session.title}" created successfully!')
            return redirect('employee_app:batch_detail', pk=session.batch.pk)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = SessionForm(initial=initial)

    return render(request, 'employee_app/add_session.html', {'form': form, 'batch_id': batch_id})

from django.http import JsonResponse

@login_required
def get_batch_trainers(request):
    batch_id = request.GET.get('batch_id')
    if not batch_id:
        return JsonResponse({'trainers': []})

    try:
        batch = Batch.objects.get(pk=batch_id)
        trainers = [
            {'id': trainer.id, 'name': trainer.full_name or trainer.email}
            for trainer in batch.trainers.all()
        ]
        return JsonResponse({'trainers': trainers})
    except Batch.DoesNotExist:
        return JsonResponse({'trainers': []})


@login_required
def edit_session(request, pk):
    if not has_permission(request.user, 'training', 'edit'):
        return HttpResponseForbidden("You don't have permission to edit sessions.")

    session = get_object_or_404(Session, pk=pk)

    if request.method == 'POST':
        form = SessionEditForm(request.POST, instance=session)
        if form.is_valid():
            form.save()
            messages.success(request, f'Session "{session.title}" updated successfully!')
            return redirect('employee_app:session_detail', pk=session.pk)
    else:
        form = SessionEditForm(instance=session)

    return render(request, 'employee_app/edit_session.html', {
        'form': form,
        'session': session
    })


# views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from .models import Session, Attendance, Batch
from .forms import SessionForm, SessionEditForm  # if you have them

@login_required
def session_detail(request, pk):
    session = get_object_or_404(Session, pk=pk)

    # Check permission to view (trainer/admin or enrolled employee)
    can_view = (
        request.user in session.batch.trainers.all() or
        request.user.role in ['admin', 'super_admin'] or
        request.user in session.batch.employees.all()
    )
    if not can_view:
        return HttpResponseForbidden("You don't have permission to view this session.")

    # Can mark attendance? (trainers of batch or admin)
    can_mark_attendance = (
        request.user in session.batch.trainers.all() or
        request.user.role in ['admin', 'super_admin']
    )

    context = {
        'session': session,
        'batch': session.batch,
        'attendance_records': session.attendance_records.all() if session.attendance_taken else None,
        'can_mark_attendance': can_mark_attendance,
        'attendance_percentage': session.attendance_percentage if hasattr(session, 'attendance_percentage') else 0,
    }

    return render(request, 'employee_app/session_detail.html', context)


@login_required
def mark_attendance(request, session_id):
    session = get_object_or_404(Session, pk=session_id)

    # Permission check
    if not (request.user in session.batch.trainers.all() or request.user.role in ['admin', 'super_admin']):
        return HttpResponseForbidden("You don't have permission to mark/update attendance.")

    employees = session.batch.employees.all().order_by('full_name')

    # Load existing attendance if already marked
    existing_attendance = {}
    if session.attendance_taken:
        for record in session.attendance_records.all():
            existing_attendance[record.employee.id] = {
                'status': record.status,
                'notes': record.notes,
            }

    if request.method == 'POST':
        # Delete old attendance records (clean slate)
        session.attendance_records.all().delete()

        # Create new records
        for emp in employees:
            status = request.POST.get(f'status_{emp.id}', 'absent')
            notes = request.POST.get(f'notes_{emp.id}', '').strip()

            Attendance.objects.create(
                session=session,
                employee=emp,
                status=status,
                notes=notes,
                marked_by=request.user
            )

        # Always set flag to True (even if updating)
        session.attendance_taken = True
        session.save(update_fields=['attendance_taken'])

        messages.success(request, f"Attendance successfully {'updated' if existing_attendance else 'marked'}!")
        return redirect('employee_app:session_detail', pk=session.pk)

    context = {
        'session': session,
        'employees': employees,
        'existing_attendance': existing_attendance,
        'is_update': bool(existing_attendance),
    }

    return render(request, 'employee_app/mark_attendance.html', context)

# ───────────────────────────────────────────────
# Assignment Views
# ───────────────────────────────────────────────

@login_required
def create_assignment(request):
    if not has_permission(request.user, 'training', 'create'):
        return HttpResponseForbidden("You don't have permission to create assignments.")

    if request.method == 'POST':
        form = AssignmentForm(request.POST)
        if form.is_valid():
            assignment = form.save(commit=False)
            assignment.created_by = request.user
            assignment.save()
            messages.success(request, f'Assignment "{assignment.title}" created successfully!')
            return redirect('employee_app:batch_detail', pk=assignment.batch.pk)
    else:
        form = AssignmentForm()

    return render(request, 'employee_app/add_assignment.html', {'form': form})


@login_required
def edit_assignment(request, pk):
    if not has_permission(request.user, 'training', 'edit'):
        return HttpResponseForbidden("You don't have permission to edit assignments.")

    assignment = get_object_or_404(Assignment, pk=pk)

    if request.method == 'POST':
        form = AssignmentEditForm(request.POST, instance=assignment)
        if form.is_valid():
            form.save()
            messages.success(request, f'Assignment "{assignment.title}" updated successfully!')
            return redirect('employee_app:assignment_detail', pk=assignment.pk)
    else:
        form = AssignmentEditForm(instance=assignment)

    return render(request, 'employee_app/edit_assignment.html', {
        'form': form,
        'assignment': assignment
    })


@login_required
def assignment_detail(request, pk):
    assignment = get_object_or_404(Assignment, pk=pk)

    context = {
        'assignment': assignment,
        'submissions': assignment.submissions.all(),
        'submission_count': assignment.submissions.count(),
    }

    return render(request, 'employee_app/assignment_detail.html', context)


# ───────────────────────────────────────────────
# Employee Submission View
# ───────────────────────────────────────────────

@login_required
def submit_assignment(request, assignment_id):
    assignment = get_object_or_404(Assignment, pk=assignment_id)

    if request.user.role != 'employee':
        return HttpResponseForbidden("Only employees can submit assignments.")

    # Check if already submitted
    submission, created = Submission.objects.get_or_create(
        assignment=assignment,
        employee=request.user,
        defaults={'submitted_at': timezone.now()}
    )

    if request.method == 'POST':
        form = SubmissionForm(request.POST, request.FILES, instance=submission)
        if form.is_valid():
            form.save()
            messages.success(request, 'Assignment submitted successfully!')
            return redirect('employee_app:my_training')
    else:
        form = SubmissionForm(instance=submission)

    return render(request, 'employee_app/assignment_submit.html', {
        'form': form,
        'assignment': assignment,
        'submission': submission
    })

